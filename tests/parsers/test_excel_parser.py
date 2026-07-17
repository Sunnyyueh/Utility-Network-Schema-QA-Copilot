from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from un_schema_qa.models.enums import InputFormat
from un_schema_qa.parsers.base import InputParseError
from un_schema_qa.parsers.excel_parser import ExcelParser


def test_excel_parser_reads_all_sheets_in_stable_order(tmp_path: Path) -> None:
    path = tmp_path / "schema.xlsx"
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame([{"name": "PVC"}]).to_excel(
            writer,
            sheet_name="Domains",
            index=False,
        )
        pd.DataFrame([{"name": "Material", "length": None}]).to_excel(
            writer,
            sheet_name="Fields",
            index=False,
        )

    document = ExcelParser().parse(path)

    assert document.input_format is InputFormat.XLSX
    assert [sheet.name for sheet in document.sheets] == ["Domains", "Fields"]
    assert document.sheets[1].rows[0]["length"] is None


def test_excel_parser_preserves_path_and_sorts_sheet_names(tmp_path: Path) -> None:
    path = tmp_path / "schema.xlsx"
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame([{"name": "Material"}]).to_excel(
            writer,
            sheet_name="Fields",
            index=False,
        )
        pd.DataFrame([{"name": "PVC"}]).to_excel(
            writer,
            sheet_name="Domains",
            index=False,
        )

    document = ExcelParser().parse(path)

    assert document.path == str(path)
    assert document.input_format is InputFormat.XLSX
    assert [sheet.name for sheet in document.sheets] == ["Domains", "Fields"]


def test_excel_parser_serializes_missing_cells_as_json_null(tmp_path: Path) -> None:
    path = tmp_path / "schema.xlsx"
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame([{"name": "Material", "length": None}]).to_excel(
            writer,
            sheet_name="Fields",
            index=False,
        )

    document = ExcelParser().parse(path)

    assert '"length":null' in document.model_dump_json()


def test_excel_parser_preserves_header_only_sheet_as_empty_rows(tmp_path: Path) -> None:
    path = tmp_path / "schema.xlsx"
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame(columns=["name", "length"]).to_excel(
            writer,
            sheet_name="Fields",
            index=False,
        )

    document = ExcelParser().parse(path)

    assert document.sheets[0].name == "Fields"
    assert document.sheets[0].rows == ()


@pytest.mark.parametrize(
    ("contents", "cause_type"),
    [(None, FileNotFoundError), (b"not an Excel workbook", ValueError)],
    ids=["missing", "corrupt"],
)
def test_excel_parser_wraps_workbook_errors(
    tmp_path: Path,
    contents: bytes | None,
    cause_type: type[BaseException],
) -> None:
    path = tmp_path / "schema.xlsx"
    if contents is not None:
        path.write_bytes(contents)

    with pytest.raises(InputParseError) as exc_info:
        ExcelParser().parse(path)

    error = exc_info.value
    assert error.code == "XLSX_PARSE_FAILED"
    assert error.path == path
    assert error.detail
    assert isinstance(error.__cause__, cause_type)
    assert error.detail == str(error.__cause__)


def test_excel_parser_reports_non_json_cells_as_invalid_rows(tmp_path: Path) -> None:
    path = tmp_path / "schema.xlsx"
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame([{"observed_at": datetime(2026, 7, 17, 12, 0)}]).to_excel(
            writer,
            sheet_name="Fields",
            index=False,
        )

    with pytest.raises(InputParseError) as exc_info:
        ExcelParser().parse(path)

    error = exc_info.value
    assert error.code == "ROWS_INVALID"
    assert error.path == path


def test_excel_parser_does_not_modify_input_bytes(tmp_path: Path) -> None:
    path = tmp_path / "schema.xlsx"
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame([{"name": "Material"}]).to_excel(
            writer,
            sheet_name="Fields",
            index=False,
        )
    original = path.read_bytes()

    ExcelParser().parse(path)

    assert path.read_bytes() == original


@pytest.mark.parametrize(
    ("headers", "duplicate"),
    [(["name", "name"], "name"), ([1, "1"], "1")],
    ids=["identical", "normalized-output-key"],
)
def test_excel_parser_rejects_duplicate_normalized_headers(
    tmp_path: Path,
    headers: list[object],
    duplicate: str,
) -> None:
    path = tmp_path / "schema.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Fields"
    worksheet.append(headers)
    worksheet.append(["Material", "PVC"])
    workbook.save(path)

    with pytest.raises(InputParseError) as exc_info:
        ExcelParser().parse(path)

    error = exc_info.value
    assert error.code == "XLSX_PARSE_FAILED"
    assert error.path == path
    assert error.detail == f"worksheet 'Fields': duplicate header {duplicate!r}"


def test_excel_parser_rejects_blank_or_missing_headers(tmp_path: Path) -> None:
    path = tmp_path / "schema.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Fields"
    worksheet.append(["name", None])
    worksheet.append(["Material", "PVC"])
    workbook.save(path)

    with pytest.raises(InputParseError) as exc_info:
        ExcelParser().parse(path)

    error = exc_info.value
    assert error.code == "XLSX_PARSE_FAILED"
    assert error.path == path
    assert error.detail == "worksheet 'Fields': blank header at column 2"


def test_excel_parser_rejects_a_missing_header_row(tmp_path: Path) -> None:
    path = tmp_path / "schema.xlsx"
    workbook = Workbook()
    workbook.active.title = "Fields"
    workbook.save(path)

    with pytest.raises(InputParseError) as exc_info:
        ExcelParser().parse(path)

    error = exc_info.value
    assert error.code == "XLSX_PARSE_FAILED"
    assert error.path == path
    assert error.detail == "worksheet 'Fields': missing header row"
